import unittest
from openpyxl import load_workbook


class ParseExcelTestCase(unittest.TestCase):

	def test_parse_excel(self):

		if __name__ == '__main__':

			wb = load_workbook('test.xlsx')

			print(f'work sheetnames {wb.worksheets}')

			sheet = wb.active

			print(f'sheet names {wb.sheetnames}')

			print(f'tables {sheet.tables}')

			for col in sheet.columns:
				print(f'col: {col}')

			print(f'Accessing directly')
			print(f'A1 {sheet["A1"].value}')
			print(f'B1 {sheet["B1"].value}')
			print(f'C1 {sheet["C1"].value}')
			print(f'D1 {sheet["D1"].value}')
			print(f'A2 {sheet["A2"].value}')
			print(f'B2 {sheet["B2"].value}')
			print(f'C2 {sheet["C2"].value}')
			print(f'D2 {sheet["D2"].value}')
			print(f'A3 {sheet["A3"].value}')
			print(f'B3 {sheet["B3"].value}')
			print(f'C3 {sheet["C3"].value}')
			print(f'D3 {sheet["D3"].value}')

			print(f'Accessing by cell')
			print(f'header {sheet.cell(row=1, column=1).value}')

			for i in range(2, 5):
				for j in range(1, 5):
					print(f'{i}:{j} {sheet.cell(row=i, column=j).value}')

			print(f'Accessing by slice')
			# Access all cells from A1 to D11
			for row in sheet["A1:D2"]:
				print([x.value for x in row])

			print(f'\ndone')

		self.assertEqual(True, True)


if __name__ == '__main__':
	unittest.main()
