import argparse
from helper.base.logger import log
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.read_only import ReadOnlyCell
import os
from typing import Protocol
from helper import EnhancedWriter

__version__ = "0.0.2"


class Args(Protocol):
    file: str
    lowercase: bool


class DataRow:
    def __init__(self, data) -> None:
        self.data = data

    def __getitem__(self, key):
        return self.data[key]

    @property
    def table_schema(self) -> str:
        """表schema名"""
        return self.data[0]

    @property
    def table_name(self) -> str:
        """表名"""
        return self.data[2]

    @property
    def table_cn_name(self) -> str:
        """表中文名称"""
        return self.data[1]

    @property
    def column_name(self) -> str:
        """列名"""
        return self.data[4]

    @property
    def column_cn_name(self) -> str:
        """中文列名"""
        return self.data[3]

    @property
    def column_type(self) -> str:
        """映射规则"""
        return self.data[5]

    @property
    def column_is_pk(self) -> bool:
        """是否主键"""
        return self.data[6] == "Y"

    @property
    def column_null_str(self) -> str:
        """是否非空字符串"""
        if self.data[7] == "Y":
            return "NOT NULL"
        return "NULL"

    @property
    def column_default_value(self) -> str:
        """默认值"""
        return self.data[11]


class DataGroup:
    def __init__(self, group_id, title) -> None:
        self.data: list[DataRow] = []
        self.group_id = group_id
        self.title = title

    def get_primary_keys(self):
        keys = filter(lambda x: x.column_is_pk, self.data)
        return list(map(lambda x: x.column_name, keys))

    def write_sql(self, w: EnhancedWriter):
        first_row = self.data[0]
        w.writeln(
            f"-- Table: {first_row.table_schema}.{first_row.table_name} {self.title}"
        )
        pks = self.get_primary_keys()
        w.writeln(f"CREATE TABLE {first_row.table_schema}.{first_row.table_name} (")
        for i in range(len(self.data)):
            d = self.data[i]
            w.write(f"    {d.column_name} {d.column_type} {d.column_null_str}")
            if d.column_default_value is not None:
                w.write(f" DEFAULT {d.column_default_value}")
            if i < len(self.data) - 1 or len(pks) > 0:
                w.write(",")
            w.writeln(f"    -- {d.column_cn_name}")
        if len(pks) > 0:
            w.writeln(
                f"    CONSTRAINT pk_{self.group_id} PRIMARY KEY ({','.join(pks)})"
            )

        w.writeln(f");")

        w.writeln(
            f"COMMENT ON TABLE {first_row.table_schema}.{first_row.table_name} IS '{self.title}';"
        )
        for i in range(len(self.data)):
            d = self.data[i]
            w.writeln(
                f"COMMENT ON COLUMN {d.table_schema}.{d.table_name}.{d.column_name} IS '{d.column_cn_name}';"
            )


class ModelFile:
    def __init__(self, sheet: Worksheet) -> None:

        # self.source_tables = sheet[1]
        # self.target_table_name = sheet["B4"].value
        # self.target_table_chinese_name = sheet["B3"].value
        # self.description = sheet["B5"].value
        # self.target_table_type = None
        # self.load_type = None
        # self.logic_increment = None

        self.sheet_name = sheet.title
        groups = {}

        for v in sheet.iter_rows(min_row=3, values_only=True):
            if v[2] is not None:
                gid = v[2]
                if gid not in groups:
                    groups[gid] = DataGroup(gid, sheet.title)
                group: DataGroup = groups[gid]
                group.data.append(DataRow(v))

        self.data: list[DataGroup] = list(groups.values())

    def generate_ddl_file(self, writer: EnhancedWriter, args: Args):
        for group in self.data:
            if group.group_id is None:
                continue
            group.write_sql(writer)
            writer.writeln()
            log.info(f"生成DDL脚本:{group.group_id}")


def find_model_file():
    cwd = os.getcwd()
    entrys = os.listdir(cwd)

    for entry in entrys:
        file = os.path.join(cwd, entry)
        if (
            os.path.isfile(file)
            and file.endswith(".xlsx")
            and not entry.startswith("~$")
        ):
            return file
    return None


def is_model_sheet(sheet: Worksheet):
    v: ReadOnlyCell = sheet["A2"]
    if v.value == "Schema":
        return True
    return False


def load_model_file(args: Args):

    book: Workbook = load_workbook(
        args.file, read_only=True, data_only=True, keep_vba=False
    )

    for sheet_name in book.sheetnames:
        sheet: Worksheet = book[sheet_name]
        if is_model_sheet(sheet):
            model = ModelFile(sheet)
            yield model


def main():

    parser = argparse.ArgumentParser(description="DDL脚本生成程序")
    parser.add_argument(
        "--version", action="version", version=__version__, help="显示程序版本号"
    )
    parser.add_argument("file", type=str, nargs="?", default=None, help="Excel字典文件")
    parser.add_argument("--lowercase", action="store_false", help="将表名和字段名强制转换为小写")

    args: Args = parser.parse_args()

    log.info(f"{parser.description} {__version__}")

    if args.file is None:
        args.file = find_model_file()

    if args.file is None:
        log.error("未找到Excel字典文件")
        exit(1)

    file = f"{args.file}.sql"
    with open(file, mode="w", encoding="utf-8") as writer:
        e_writer = EnhancedWriter(writer)
        for model in load_model_file(args):
            model.generate_ddl_file(e_writer, args)


if __name__ == "__main__":
    main()
