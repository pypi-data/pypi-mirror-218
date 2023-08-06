from re import compile, IGNORECASE
from io import BytesIO
import time
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import requests
from .juniper import get_domain_for as jnpr_domain_for
from .countries import get_country_regex

"""
This module reads a spreadsheet from cisco.com, which is what is houses the data being used on:
https://www.cisco.com/c/dam/assets/prod/wireless/wireless-compliance-tool/index.html
This page is basically javascript, reading data from the spreadsheet, and then offering it in an interactive web form.
This is all nice and fine, but in certain cases it would be neat if you could automate which specific AP model to buy,
given the country.
"""
compliance_url = "https://www.cisco.com/c/dam/assets/prod/wireless/wireless-compliance-tool/ComplianceStatus.xls"
platforms = {
    "Indoor": (2, 4),
    "Outdoor and Industrial": (2, 3),
}
_cache = {}
model_pat = compile(r"^(AIR-|IW|C[9W]).*", IGNORECASE)
MAX_COL = 160


def parse_row(row: int, cols: dict[str, int], sheet: Worksheet) -> dict[str, str]:
    """
    parse a row into a dict
    :param int row: row index
    :param dict cols: dict of header, column index
    :param Worksheet sheet: sheet to parse data from
    :return: dict of values key'ed by their column name
    :rtype: dict[str, str]
    """
    vals = {}
    for header, col in cols.items():
        cell = sheet.cell(row, col)
        vals[header] = cell.value
    return vals


def get_book() -> Workbook:
    """
    Lazily fetch the spreadsheet, fresh from cisco.com
    :return:
    :rtype: Workbook
    """
    global _cache
    if book := _cache.get("wb", None):
        return book

    xls_data = requests.get(compliance_url, allow_redirects=True).content
    book = load_workbook(BytesIO(xls_data), read_only=True, data_only=True)
    print(f"WB loaded and parsed: {time.time():.2f}")
    _cache["wb"] = book
    return book


def get_models_by_platform() -> dict[str, list[str]]:
    """
    Retrieve all models, arranged as a dict of platforms, each having a list of models
    :return:
    :rtype: dict[str, list]
    """
    global _cache
    if mbp := _cache.get("mbp"):
        return mbp
    models_by_platform = {}
    for p, indexes in platforms.items():
        models_by_platform[p] = set()
        sheet = get_platform_sheet(p)
        start_col = indexes[1] + 1
        try:
            for col in range(start_col, MAX_COL):
                txt = sheet.cell(1, col).value
                if model_pat.search(txt):
                    models_by_platform[p].add(txt.upper())
                else:
                    break
        except IndexError:
            pass
    print(f"MbP loaded and parsed: {time.time():.2f}")
    _cache["mbp"] = models_by_platform
    return models_by_platform


def get_models() -> list[str]:
    """
    Get all models in a list
    :return:
    :rtype: list[str]
    """
    global _cache
    if models := _cache.get("models"):
        return models
    models = set()
    for platform, platform_models in get_models_by_platform().items():
        for model in platform_models:
            models.add(model)
    _cache["models"] = list(models)
    return _cache["models"]


def get_platform_sheet(platform: str) -> Worksheet:
    """
    :param str platform:
    :rtype: Worksheet
    """
    wb = get_book()
    try:
        return wb[platform]
    except Exception as sheet_error:
        if "&" not in platform:
            raise sheet_error
        return wb[platform.replace("&", "and")]


def get_domain_for(model: str, country=None) -> list[str]:
    """
    Get all valid domains for a given model, in a particular country.
    :param str model:
    :param str country:
    :return:
    """
    valid_domains = set()
    found_country = False
    if country:
        cpat = get_country_regex(country)

    for platform in get_platform_for(model):
        sheet = get_platform_sheet(platform)
        h = {
            "Country": platforms[platform][0],
            "Regulatory Domain": platforms[platform][1],
        }
        for header, col in h.items():
            if header != sheet.cell(1, col).value:
                raise ValueError(
                    "Found unexpected header at 1,{}: {} vs {}".format(
                        col, sheet.cell(1, col).value, header
                    )
                )
        if col := _cache.get(f"col_{model.upper()}"):
            h[model] = col
        else:
            # models start just after regulatory domain column
            for col in range(platforms[platform][1] + 1, MAX_COL):
                try:
                    if cell_model := sheet.cell(1, col).value:
                        upper_model = cell_model.upper()
                        _cache.setdefault(f"col_{upper_model}", col)
                        if upper_model == model.upper():
                            h[model] = col
                except IndexError:
                    if model not in h:
                        msg = f"Unable to find the model {model} in the {platform} platform?"
                        raise ValueError(msg)
        if countries := _cache.get(f"countries_{sheet.title}"):
            pass
        else:
            countries = []
            for row_no in range(4, sheet.max_row + 1):
                row = parse_row(row_no, h, sheet)
                if not row or (not row["Country"] and not row["Regulatory Domain"]):
                    continue
                countries.append(row)
            _cache[f"countries_{sheet.title}"] = countries

        for row in countries:
            # this way we also match United States in United States of America
            if not country or row["Country"] and cpat.search(row["Country"]):
                found_country = True
                if row[model] == "x":
                    valid_domains.add(row["Regulatory Domain"])

    if valid_domains:
        return list(valid_domains)
    elif found_country:
        raise ValueError(
            f"Found {model} for {country} - but no active regulatory domains?"
        )
    elif country:
        rd = jnpr_domain_for(country)
        return [rd]
    else:
        raise ValueError(f"Couldn't find any country matching {country}")


def get_country_models(model: str) -> list[str]:
    """
    Get all valid domain-specific models for a given model.
    :param str model:
    :return:
    """
    return get_models_for(model)


def get_models_for(model: str, country=None) -> list[str]:
    """
    Get precise model name for a model and country
    :param str model:
    :param str country:
    :return:
    :rtype: list[str]
    """

    domains = get_domain_for(model, country)
    if not domains:
        raise ValueError(
            f"Unable to find any valid regulatory domains for {model} in {country}"
        )

    return [f"{model.upper()}{dom}{'-K9' if model.upper().startswith('AIR-') else ''}" for dom in domains]


def get_platform_for(model: str) -> list[str]:
    """
    Fetch platform and indexes for respectively country and regulatory domain as a tuple
    :param model:
    :return:
    :rtype: (str, int, int)
    """
    global _cache
    if platform := _cache.get(f"platform_{model}"):
        return platform
    models = get_models_by_platform()
    model_platforms = []
    for platform, platform_models in models.items():
        if model in platform_models:
            model_platforms.append(platform)
    if not model_platforms:
        raise ValueError(f"Couldn't find the model {model}")
    _cache[f"platform_{model}"] = model_platforms
    return model_platforms
